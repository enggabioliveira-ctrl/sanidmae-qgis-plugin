# -*- coding: utf-8 -*-
"""
Exportacao dos relatorios (trechos criticos e diagnostico geral) em Excel
(.xlsx, com formatacao) quando openpyxl estiver disponivel, com fallback
para CSV (com virgula decimal, padrao Brasil) quando nao estiver.

Parte do RedeCascataDockWidget, dividido em mixins por assunto so pra
organizacao/leitura - continua sendo tudo a mesma classe/instancia em tempo
de execucao (ver rede_cascata_dockwidget.py, onde os mixins sao combinados).
"""
import csv
import os
from datetime import datetime

from qgis.core import (
    QgsProject,
    QgsField,
    QgsFeatureRequest,
    QgsSymbol,
    QgsRendererCategory,
    QgsCategorizedSymbolRenderer,
    QgsPalLayerSettings,
    QgsTextFormat,
    QgsTextBackgroundSettings,
    QgsVectorLayerSimpleLabeling,
    QgsProperty,
    QgsWkbTypes,
    QgsLineSymbol,
    QgsMarkerSymbol,
    QgsMarkerLineSymbolLayer,
    QgsSymbolLayer,
    QgsCoordinateTransform,
    QgsPrintLayout,
    QgsLayoutItemMap,
    QgsLayoutItemLegend,
    QgsLayoutItemScaleBar,
    QgsLayoutItemLabel,
    QgsLayoutPoint,
    QgsUnitTypes,
)
from qgis.PyQt.QtCore import Qt, QVariant, QEventLoop, QRectF
from qgis.PyQt.QtGui import QColor, QFont
from qgis.PyQt.QtWidgets import (
    QDockWidget,
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QFormLayout,
    QLabel,
    QComboBox,
    QPushButton,
    QLineEdit,
    QGroupBox,
    QFileDialog,
    QMessageBox,
    QDialog,
    QDialogButtonBox,
    QDoubleSpinBox,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QScrollArea,
    QSizePolicy,
    QInputDialog,
    QPlainTextEdit,
    QCheckBox,
)

from .db import RedeDB
from .calculo import calcular_tudo, calcular_vazao_bacia, ErroTopologia
from .maptool_link import VincularMapTool
from .geo_utils import calcular_area_ha, calcular_comprimento_m, nova_distance_area
from .collapsible import GrupoRecolhivel
from .constantes import CAMPOS_RESULTADO, CAMPOS_BACIA


class RelatoriosMixin:
    def _tentar_openpyxl(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            return {
                "openpyxl": openpyxl,
                "Font": Font,
                "PatternFill": PatternFill,
                "Alignment": Alignment,
                "get_column_letter": get_column_letter,
            }
        except ImportError:
            return None

    def _numero_br(self, valor, casas=3):
        """Formata numero com virgula decimal (padrao Brasil) - usado so no
        CSV de reserva, quando o openpyxl nao esta disponivel. Num arquivo
        CSV com ';' como separador de coluna, o Excel-BR espera numero com
        virgula; escrever com ponto faz ele ler tudo errado (ex: 3.406
        virando 3406)."""
        if valor is None:
            return ""
        return f"{valor:,.{casas}f}".replace(",", "_").replace(".", ",").replace("_", ".")

    def _ajustar_largura_colunas(self, planilha, ferramentas, num_colunas, limite_linhas=300):
        """Estima a largura das colunas olhando so uma amostra das
        primeiras linhas (nao a planilha inteira) - com muitas linhas
        (rede grande, milhares de trechos), escanear tudo pra so ajustar
        largura de coluna era a maior causa da travada ao gerar o excel."""
        get_column_letter = ferramentas["get_column_letter"]
        max_linha = min(planilha.max_row, limite_linhas)
        for col in range(1, num_colunas + 1):
            letra = get_column_letter(col)
            maior = len(str(planilha.cell(row=1, column=col).value or ""))
            for linha in range(2, max_linha + 1):
                valor = planilha.cell(row=linha, column=col).value
                if valor is not None:
                    maior = max(maior, len(str(valor)))
            planilha.column_dimensions[letra].width = min(max(maior + 2, 10), 40)

    def _exportar_relatorio(self):
        if self.db is None:
            QMessageBox.warning(self, "Aviso", "Calcule antes de exportar.")
            return
        resultados = self.db.get_resultados()
        criticos = [r for r in resultados if r["critico"]]
        if not criticos:
            QMessageBox.information(self, "Info", "Nenhum trecho critico encontrado.")
            return

        ferramentas = self._tentar_openpyxl()
        criticos_ordenados = sorted(criticos, key=lambda x: x["coletor_id"])

        if ferramentas:
            caminho, _ = QFileDialog.getSaveFileName(
                self, "Salvar relatorio", "relatorio_trechos_criticos.xlsx", "Excel (*.xlsx)"
            )
            if not caminho:
                return
            self._escrever_xlsx_trechos_criticos(caminho, criticos_ordenados, ferramentas)
        else:
            caminho, _ = QFileDialog.getSaveFileName(
                self, "Salvar relatorio", "relatorio_trechos_criticos.csv", "CSV (*.csv)"
            )
            if not caminho:
                return
            self._escrever_csv_trechos_criticos(caminho, criticos_ordenados)
            QMessageBox.information(
                self,
                "Info",
                "O pacote 'openpyxl' nao esta instalado no Python do QGIS, entao o "
                "relatorio foi gerado em CSV (abre normalmente no Excel, com virgula "
                "decimal). Se quiser o arquivo em .xlsx com formatacao (cores, "
                "largura de coluna automatica), peca pra alguem instalar o openpyxl "
                "no ambiente Python do QGIS (OSGeo4W Shell: "
                "python -m pip install openpyxl).",
            )
            return

        QMessageBox.information(self, "OK", f"Relatorio exportado: {caminho}")

    def _escrever_csv_trechos_criticos(self, caminho, criticos):
        with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f, delimiter=";")
            writer.writerow(
                [
                    "coletor_id",
                    "vazao_acumulada_ls",
                    "declividade_usada_pct",
                    "dn_calculado_m",
                    "dn_existente_m",
                    "diferenca_m",
                ]
            )
            for r in criticos:
                dif = (r["dn_calculado"] or 0) - (r["dn_existente"] or 0)
                s_pct = r.get("inclinacao_usada")
                writer.writerow(
                    [
                        r["coletor_id"],
                        self._numero_br(r["vazao_acumulada"]),
                        self._numero_br(s_pct * 100 if s_pct is not None else None),
                        self._numero_br(r["dn_calculado"]),
                        self._numero_br(r["dn_existente"]),
                        self._numero_br(dif),
                    ]
                )

    def _escrever_xlsx_trechos_criticos(self, caminho, criticos, ferramentas):
        openpyxl = ferramentas["openpyxl"]
        Font = ferramentas["Font"]
        PatternFill = ferramentas["PatternFill"]
        Alignment = ferramentas["Alignment"]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trechos criticos"

        # cria os estilos UMA vez e reaproveita - criar um objeto de estilo
        # novo pra cada celula/linha e o principal motivo de travar com
        # muitas linhas (rede grande, milhares de trechos)
        fonte_cabecalho = Font(bold=True, color="FFFFFF")
        fundo_cabecalho = PatternFill("solid", fgColor="C0392B")
        alinhamento_centro = Alignment(horizontal="center")
        fundo_critico = PatternFill("solid", fgColor="FADBD8")

        cabecalho = [
            "Coletor", "Vazao acumulada (L/s)", "Declividade usada (%)",
            "DN calculado (m)", "DN existente (m)", "Diferenca (m)",
        ]
        ws.append(cabecalho)
        for celula in ws[1]:
            celula.font = fonte_cabecalho
            celula.fill = fundo_cabecalho
            celula.alignment = alinhamento_centro

        for r in criticos:
            dif = (r["dn_calculado"] or 0) - (r["dn_existente"] or 0)
            s_pct = r.get("inclinacao_usada")
            linha = [
                r["coletor_id"],
                round(r["vazao_acumulada"], 3),
                round(s_pct * 100, 3) if s_pct is not None else None,
                round(r["dn_calculado"], 3) if r["dn_calculado"] is not None else None,
                round(r["dn_existente"], 3) if r["dn_existente"] is not None else None,
                round(dif, 3),
            ]
            ws.append(linha)
            for celula in ws[ws.max_row]:
                celula.fill = fundo_critico

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        self._ajustar_largura_colunas(ws, ferramentas, len(cabecalho))
        wb.save(caminho)

    # ------------------------------------------------- diagnostico geral ----
    def _gerar_relatorio_diagnostico(self):
        if self.db is None:
            QMessageBox.warning(self, "Aviso", "Crie/abra o banco de calculo e calcule antes de gerar o relatorio.")
            return
        resultados = self.db.get_resultados()
        if not resultados:
            QMessageBox.information(self, "Info", "Nenhum resultado calculado ainda. Clique em 'Calcular tudo' primeiro.")
            return

        ferramentas = self._tentar_openpyxl()

        if ferramentas:
            caminho, _ = QFileDialog.getSaveFileName(
                self, "Salvar diagnostico geral", "diagnostico_geral_rede.xlsx", "Excel (*.xlsx)"
            )
            if not caminho:
                return
            self._escrever_xlsx_diagnostico(caminho, resultados, ferramentas)
        else:
            caminho, _ = QFileDialog.getSaveFileName(
                self, "Salvar diagnostico geral", "diagnostico_geral_rede.csv", "CSV (*.csv)"
            )
            if not caminho:
                return
            self._escrever_csv_diagnostico(caminho, resultados)
            QMessageBox.information(
                self,
                "Info",
                "O pacote 'openpyxl' nao esta instalado no Python do QGIS, entao o "
                "relatorio foi gerado em CSV (abre normalmente no Excel, com virgula "
                "decimal). Se quiser o arquivo em .xlsx com abas separadas e "
                "formatacao, peca pra alguem instalar o openpyxl no ambiente Python "
                "do QGIS (OSGeo4W Shell: python -m pip install openpyxl).",
            )
            return

        QMessageBox.information(self, "OK", f"Diagnostico geral exportado: {caminho}")

    def _dados_diagnostico(self, resultados):
        """Monta os dados comuns usados tanto no xlsx quanto no CSV de
        diagnostico geral, pra nao duplicar a logica duas vezes."""
        parametros = self.db.get_parametros()
        excecoes = self.db.get_all_excecoes()
        bacia_dados = self.db.get_all_bacia_dados()
        bacia_coletor_map = self.db.get_bacia_coletor_map()
        coletor_destino_map = self.db.get_coletor_destino_map()

        qf = float(parametros.get("qf", 150))
        C = float(parametros.get("C", 0.8))
        k1 = float(parametros.get("k1", 1.2))
        k2 = float(parametros.get("k2", 1.5))

        linhas_bacias = []
        da = nova_distance_area(self.layer_bacias) if self.layer_bacias else None
        for feicao in (self.layer_bacias.getFeatures() if self.layer_bacias else []):
            bacia_id = str(feicao.attribute(self.campo_id_bacia))
            info = bacia_dados.get(bacia_id)
            if info is None or info.get("densidade_hab_ha") is None:
                continue
            if info.get("area_ha_manual") is not None:
                area_ha = info["area_ha_manual"]
            else:
                area_ha = calcular_area_ha(self.layer_bacias, feicao.geometry(), da)
            densidade = info["densidade_hab_ha"]
            hab, q = calcular_vazao_bacia(area_ha, densidade, qf, C, k1, k2)
            linhas_bacias.append(
                (bacia_id, bacia_coletor_map.get(bacia_id, ""), area_ha, densidade, hab, q)
            )

        criticos = [r for r in resultados if r["critico"]]

        return {
            "parametros": parametros,
            "excecoes": excecoes,
            "bacia_dados": bacia_dados,
            "coletor_destino_map": coletor_destino_map,
            "linhas_bacias": linhas_bacias,
            "criticos": criticos,
        }

    def _escrever_csv_diagnostico(self, caminho, resultados):
        dados = self._dados_diagnostico(resultados)
        parametros = dados["parametros"]
        excecoes = dados["excecoes"]

        rotulos_parametros = {
            "qf": "Consumo per capita qf (L/hab.dia)",
            "C": "Coeficiente de retorno C",
            "k1": "Coef. dia de maior consumo k1",
            "k2": "Coef. hora de maior consumo k2",
            "inclinacao": "Inclinacao S padrao (m/m)",
            "lamina_relativa": "Lamina relativa f(h/D) padrao",
            "rugosidade": "Rugosidade de Manning n padrao",
            "formula_diametro": "Formula do diametro calculado",
        }

        with open(caminho, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f, delimiter=";")

            w.writerow(["DIAGNOSTICO GERAL - DIMENSIONAMENTO DE REDE DE ESGOTO"])
            w.writerow(["Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")])
            w.writerow(["Camada de bacias", self.layer_bacias.name() if self.layer_bacias else ""])
            w.writerow(["Camada de coletores", self.layer_coletores.name() if self.layer_coletores else ""])
            w.writerow(["Banco de calculo", self.db.path])
            w.writerow([])

            w.writerow(["=== PARAMETROS GLOBAIS ADOTADOS ==="])
            for chave, rotulo in rotulos_parametros.items():
                valor = parametros.get(chave, "")
                w.writerow([rotulo, self._numero_br(float(valor), 4) if chave != "formula_diametro" else valor])
                if chave == "inclinacao":
                    try:
                        pct = float(valor) * 100
                        w.writerow(["Inclinacao S padrao (%)", self._numero_br(pct, 4)])
                    except (TypeError, ValueError):
                        pass
            w.writerow([])

            w.writerow(["=== EXCECOES POR TRECHO (sobrescrevem o parametro global) ==="])
            if excecoes:
                w.writerow(["coletor_id", "parametro", "valor_adotado", "valor_em_%_se_inclinacao"])
                for coletor_id in sorted(excecoes.keys()):
                    for parametro, valor in excecoes[coletor_id].items():
                        pct_txt = ""
                        if parametro == "inclinacao":
                            try:
                                pct_txt = self._numero_br(float(valor) * 100, 4)
                            except (TypeError, ValueError):
                                pct_txt = ""
                        w.writerow([coletor_id, parametro, self._numero_br(float(valor), 4), pct_txt])
            else:
                w.writerow(["(nenhuma excecao cadastrada - todos os trechos usam os parametros globais)"])
            w.writerow([])

            w.writerow(["=== DENSIDADES ADOTADAS POR BACIA ==="])
            w.writerow(["bacia_id", "coletor_destino", "area_ha", "densidade_hab_ha", "populacao_estimada", "vazao_propria_ls"])
            for bacia_id, destino, area_ha, densidade, hab, q in dados["linhas_bacias"]:
                w.writerow([
                    bacia_id, destino, self._numero_br(area_ha), densidade,
                    self._numero_br(hab, 0), self._numero_br(q),
                ])
            w.writerow([])

            w.writerow(["=== CASCATA COLETOR -> COLETOR DE JUSANTE ==="])
            coletor_destino_map = dados["coletor_destino_map"]
            if coletor_destino_map:
                w.writerow(["coletor_id", "coletor_destino_id"])
                for coletor_id in sorted(coletor_destino_map.keys()):
                    w.writerow([coletor_id, coletor_destino_map[coletor_id]])
            else:
                w.writerow(["(nenhum vinculo coletor->coletor cadastrado)"])
            w.writerow([])

            w.writerow(["=== RESULTADO POR COLETOR (TODOS) ==="])
            w.writerow([
                "coletor_id", "vazao_domestica_ls", "vazao_infiltracao_ls",
                "vazao_acumulada_ls", "declividade_usada_pct",
                "dn_calculado_m", "dn_adotado_m", "dn_existente_m", "diferenca_m", "status",
                "declividade_minima_pct", "atende_declividade_minima",
                "tensao_trativa_dn_adotado_pa", "atende_tensao_trativa",
                "velocidade_final_ms", "velocidade_critica_ms", "excede_velocidade_critica",
                "lamina_reduzida_automaticamente", "comprimento_trecho_m",
                "atende_espacamento_max_pv",
            ])
            for r in sorted(resultados, key=lambda x: x["coletor_id"]):
                dn_adot = r.get("dn_adotado")
                dif = (dn_adot or 0) - (r["dn_existente"] or 0) if r["dn_existente"] is not None else None
                status = "PRECISA TROCAR/MELHORAR" if r["critico"] else "ATENDE"
                s_pct = r.get("inclinacao_usada")
                imin_pct = r.get("declividade_minima")
                ok_imin = r.get("atende_declividade_minima")
                trat_c = r.get("tensao_trativa_calc_pa")
                ok_trat_c = r.get("atende_tensao_trativa_calc")
                vf = r.get("velocidade_final_calc")
                vc = r.get("velocidade_critica_calc")
                excede_v = r.get("excede_velocidade_critica_calc")
                lam_red = r.get("lamina_reduzida_calc")
                excede_espac = r.get("excede_espacamento_pv")
                w.writerow([
                    r["coletor_id"],
                    self._numero_br(r.get("vazao_domestica")),
                    self._numero_br(r.get("vazao_infiltracao")),
                    self._numero_br(r["vazao_acumulada"]),
                    self._numero_br(s_pct * 100) if s_pct is not None else "",
                    self._numero_br(r["dn_calculado"]),
                    self._numero_br(dn_adot),
                    self._numero_br(r["dn_existente"]),
                    self._numero_br(dif),
                    status,
                    self._numero_br(imin_pct * 100, 4) if imin_pct is not None else "",
                    ("SIM" if ok_imin else "NAO") if ok_imin is not None else "",
                    self._numero_br(trat_c),
                    ("SIM" if ok_trat_c else "NAO") if ok_trat_c is not None else "",
                    self._numero_br(vf),
                    self._numero_br(vc),
                    ("SIM" if excede_v else "NAO") if excede_v is not None else "",
                    ("SIM" if lam_red else "NAO") if lam_red is not None else "",
                    self._numero_br(r.get("comprimento_trecho_m")),
                    ("NAO" if excede_espac else "SIM") if excede_espac is not None else "",
                ])
            w.writerow([])

            w.writerow(["=== RESUMO ==="])
            w.writerow(["Total de bacias com densidade definida", len(dados["bacia_dados"])])
            w.writerow(["Total de coletores calculados", len(resultados)])
            w.writerow(["Trechos criticos (precisam trocar/melhorar)", len(dados["criticos"])])
            pct = (len(dados["criticos"]) / len(resultados) * 100) if resultados else 0
            w.writerow(["Percentual de trechos criticos", self._numero_br(pct, 1) + "%"])

    def _escrever_xlsx_diagnostico(self, caminho, resultados, ferramentas):
        openpyxl = ferramentas["openpyxl"]
        Font = ferramentas["Font"]
        PatternFill = ferramentas["PatternFill"]
        Alignment = ferramentas["Alignment"]

        dados = self._dados_diagnostico(resultados)
        parametros = dados["parametros"]
        excecoes = dados["excecoes"]

        cor_cabecalho = "1F4E78"
        cor_critico = "FADBD8"

        # estilos criados uma unica vez e reaproveitados em todas as abas -
        # criar um objeto de estilo novo por celula/linha e o que mais pesa
        # com muitas linhas (rede grande, milhares de trechos)
        fonte_cabecalho = Font(bold=True, color="FFFFFF")
        fundo_cabecalho = PatternFill("solid", fgColor=cor_cabecalho)
        alinhamento_centro = Alignment(horizontal="center")
        fundo_critico = PatternFill("solid", fgColor=cor_critico)

        def nova_aba(wb, titulo, cabecalho):
            ws = wb.create_sheet(titulo)
            ws.append(cabecalho)
            for celula in ws[1]:
                celula.font = fonte_cabecalho
                celula.fill = fundo_cabecalho
                celula.alignment = alinhamento_centro
            return ws

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # ---- aba resumo ----
        ws_resumo = wb.create_sheet("Resumo")
        ws_resumo.append(["DIAGNOSTICO GERAL - DIMENSIONAMENTO DE REDE DE ESGOTO"])
        ws_resumo["A1"].font = Font(bold=True, size=14)
        linhas_resumo = [
            ("Gerado em", datetime.now().strftime("%d/%m/%Y %H:%M")),
            ("Camada de bacias", self.layer_bacias.name() if self.layer_bacias else ""),
            ("Camada de coletores", self.layer_coletores.name() if self.layer_coletores else ""),
            ("Banco de calculo", self.db.path),
            ("", ""),
            ("Total de bacias com densidade definida", len(dados["bacia_dados"])),
            ("Total de coletores calculados", len(resultados)),
            ("Trechos criticos (precisam trocar/melhorar)", len(dados["criticos"])),
            (
                "Percentual de trechos criticos",
                round((len(dados["criticos"]) / len(resultados) * 100), 1) if resultados else 0,
            ),
        ]
        for rotulo, valor in linhas_resumo:
            ws_resumo.append([rotulo, valor])
        for linha in ws_resumo.iter_rows(min_row=2, max_col=1):
            linha[0].font = Font(bold=True)
        ws_resumo.column_dimensions["A"].width = 42
        ws_resumo.column_dimensions["B"].width = 30

        # ---- aba parametros ----
        ws_param = nova_aba(wb, "Parametros globais", ["Parametro", "Valor"])
        rotulos_parametros = {
            "qf": "Consumo per capita qf (L/hab.dia)",
            "C": "Coeficiente de retorno C",
            "k1": "Coef. dia de maior consumo k1",
            "k2": "Coef. hora de maior consumo k2",
            "inclinacao": "Inclinacao S padrao (m/m)",
            "lamina_relativa": "Lamina relativa f(h/D) padrao",
            "rugosidade": "Rugosidade de Manning n padrao",
            "formula_diametro": "Formula do diametro calculado",
        }
        for chave, rotulo in rotulos_parametros.items():
            valor = parametros.get(chave, "")
            if chave == "formula_diametro":
                ws_param.append([rotulo, valor])
            else:
                try:
                    ws_param.append([rotulo, round(float(valor), 4)])
                except (TypeError, ValueError):
                    ws_param.append([rotulo, valor])
            if chave == "inclinacao":
                try:
                    ws_param.append(["Inclinacao S padrao (%)", round(float(valor) * 100, 4)])
                except (TypeError, ValueError):
                    pass
        self._ajustar_largura_colunas(ws_param, ferramentas, 2)

        # ---- aba excecoes ----
        ws_exc = nova_aba(
            wb, "Excecoes por trecho",
            ["Coletor", "Parametro", "Valor adotado", "Valor em % (se inclinacao)"],
        )
        for coletor_id in sorted(excecoes.keys()):
            for parametro, valor in excecoes[coletor_id].items():
                pct = None
                if parametro == "inclinacao":
                    try:
                        pct = round(float(valor) * 100, 4)
                    except (TypeError, ValueError):
                        pct = None
                try:
                    valor_num = round(float(valor), 4)
                except (TypeError, ValueError):
                    valor_num = valor
                ws_exc.append([coletor_id, parametro, valor_num, pct])
        self._ajustar_largura_colunas(ws_exc, ferramentas, 4)

        # ---- aba densidades por bacia ----
        ws_bac = nova_aba(
            wb, "Densidades por bacia",
            ["Bacia", "Coletor destino", "Area (ha)", "Densidade (hab/ha)", "Populacao estimada", "Vazao propria (L/s)"],
        )
        for bacia_id, destino, area_ha, densidade, hab, q in dados["linhas_bacias"]:
            ws_bac.append([bacia_id, destino, round(area_ha, 3), densidade, round(hab), round(q, 3)])
        self._ajustar_largura_colunas(ws_bac, ferramentas, 6)

        # ---- aba cascata ----
        ws_casc = nova_aba(wb, "Cascata coletor-jusante", ["Coletor", "Coletor de jusante"])
        for coletor_id in sorted(dados["coletor_destino_map"].keys()):
            ws_casc.append([coletor_id, dados["coletor_destino_map"][coletor_id]])
        self._ajustar_largura_colunas(ws_casc, ferramentas, 2)

        # ---- aba resultado por coletor (principal) ----
        ws_res = nova_aba(
            wb, "Resultado por coletor",
            ["Coletor", "Vazao domestica (L/s)", "Vazao infiltracao (L/s)",
             "Vazao acumulada (L/s)", "Declividade usada (%)", "DN calculado (m)",
             "DN adotado (m)", "DN existente (m)", "Diferenca (m)", "Status",
             "Declividade minima NBR (%)", "Atende declividade minima",
             "Tensao trativa (Pa)", "Atende tensao trativa",
             "Velocidade final (m/s)", "Velocidade critica (m/s)", "Excede velocidade critica",
             "Lamina reduzida automaticamente", "Comprimento do trecho (m)",
             "Atende espacamento max. entre PVs"],
        )
        for r in sorted(resultados, key=lambda x: x["coletor_id"]):
            dn_adot = r.get("dn_adotado")
            dif = (dn_adot or 0) - (r["dn_existente"] or 0) if r["dn_existente"] is not None else None
            status = "PRECISA TROCAR/MELHORAR" if r["critico"] else "ATENDE"
            s_pct = r.get("inclinacao_usada")
            imin_pct = r.get("declividade_minima")
            ok_imin = r.get("atende_declividade_minima")
            trat_c = r.get("tensao_trativa_calc_pa")
            ok_trat_c = r.get("atende_tensao_trativa_calc")
            vf = r.get("velocidade_final_calc")
            vc = r.get("velocidade_critica_calc")
            excede_v = r.get("excede_velocidade_critica_calc")
            lam_red = r.get("lamina_reduzida_calc")
            excede_espac = r.get("excede_espacamento_pv")
            ws_res.append([
                r["coletor_id"],
                round(r["vazao_domestica"], 3) if r.get("vazao_domestica") is not None else None,
                round(r["vazao_infiltracao"], 4) if r.get("vazao_infiltracao") is not None else None,
                round(r["vazao_acumulada"], 3),
                round(s_pct * 100, 3) if s_pct is not None else None,
                round(r["dn_calculado"], 3) if r["dn_calculado"] is not None else None,
                round(dn_adot, 3) if dn_adot is not None else None,
                round(r["dn_existente"], 3) if r["dn_existente"] is not None else None,
                round(dif, 3) if dif is not None else None,
                status,
                round(imin_pct * 100, 4) if imin_pct is not None else None,
                ("SIM" if ok_imin else "NAO") if ok_imin is not None else "",
                round(trat_c, 3) if trat_c is not None else None,
                ("SIM" if ok_trat_c else "NAO") if ok_trat_c is not None else "",
                round(vf, 3) if vf is not None else None,
                round(vc, 3) if vc is not None else None,
                ("SIM" if excede_v else "NAO") if excede_v is not None else "",
                ("SIM" if lam_red else "NAO") if lam_red is not None else "",
                round(r.get("comprimento_trecho_m"), 2) if r.get("comprimento_trecho_m") is not None else None,
                ("NAO" if excede_espac else "SIM") if excede_espac is not None else "",
            ])
            if r["critico"]:
                for celula in ws_res[ws_res.max_row]:
                    celula.fill = fundo_critico
        ws_res.freeze_panes = "A2"
        ws_res.auto_filter.ref = ws_res.dimensions
        self._ajustar_largura_colunas(ws_res, ferramentas, 20)

        wb.save(caminho)


